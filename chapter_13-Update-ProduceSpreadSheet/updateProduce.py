#! python3
#  updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl

# The produce types and their updated prices
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# Open the spreadsheet file & load the produce sheet onto the program
wb = openpyxl.load_workbook("chapter_13-Update-ProduceSpreadSheet/produceSales.xlsx")
sheet = wb["Sheet"]

# For each row, check whether the value in column A is Celery, Garlic, or Lemon
for rowNum in range(2, sheet.max_row):
    produce_name = sheet.cell(row=rowNum, column=1).value
    if produce_name in PRICE_UPDATES:
        pass

        # TODO: If it is, update the price in column B

        # TODO: Save spreadsheet to new file so we don't lose original data
